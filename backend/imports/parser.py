"""
Generic file parser — CSV and Excel (XLSX) with multi-sheet support.
Also detects One2Many child fields (headers with `/` separator).
"""
import csv
import io
from pathlib import Path

import openpyxl


def parse_file(file_obj, sheet_name=None):
    """
    Parse uploaded file (CSV or XLSX) into structured data.

    Returns:
        dict with:
            - headers: list of column header strings
            - rows: list of list of cell values (all rows)
            - sheets: list of sheet names (XLSX only, else empty)
            - child_fields: dict { relation_name: { child_headers, child_col_indices } }
    """
    filename = getattr(file_obj, 'name', '')
    ext = Path(filename).suffix.lower()

    if ext in ('.xlsx', '.xls'):
        result = _parse_xlsx(file_obj, sheet_name)
    else:
        result = _parse_csv(file_obj)

    # Detect child fields (headers with / separator)
    result['child_fields'] = _detect_child_headers(result['headers'])
    return result


def _detect_child_headers(headers):
    """
    Detect headers with '/' separator indicating One2Many child fields.

    Input: ['Vendor', 'Reference', 'order_lines/product', 'order_lines/qty']
    Output: {'order_lines': ['product', 'qty']}
    """
    child_fields = {}
    for h in headers:
        parts = h.split('/', 1)
        if len(parts) == 2:
            relation = parts[0]
            child_field = parts[1]
            if relation not in child_fields:
                child_fields[relation] = []
            child_fields[relation].append(child_field)
    return child_fields


def separate_parent_child(headers, rows, model_cls=None):
    """
    Separate flat data into parent-only and child-only structures.

    Args:
        headers: list of header strings
        rows: list of value lists
        model_cls: optional model class (to validate One2Many field names)

    Returns:
        dict:
            - parent_headers: headers without / separator
            - parent_col_indices: indices of parent columns
            - child_groups: { relation_name: { child_headers, child_col_indices, child_rows } }
    """
    child_detected = _detect_child_headers(headers)

    parent_headers = []
    parent_col_indices = []
    child_groups = {}

    for idx, h in enumerate(headers):
        parts = h.split('/', 1)
        if len(parts) == 2:
            relation, child_field = parts

            # Validate: check if relation is a valid One2Many on the model
            if model_cls:
                fd = model_cls._field_descriptors.get(relation)
                if not fd or getattr(fd, 'field_type', None) != 'one2many':
                    parent_headers.append(h)
                    parent_col_indices.append(idx)
                    continue

            if relation not in child_groups:
                child_groups[relation] = {
                    'child_headers': [],
                    'child_col_indices': [],
                    'child_field_names': [],
                    'child_rows': [],
                }
            child_groups[relation]['child_headers'].append(h)
            child_groups[relation]['child_col_indices'].append(idx)
            child_groups[relation]['child_field_names'].append(child_field)
        else:
            parent_headers.append(h)
            parent_col_indices.append(idx)

    # Extract child rows
    for relation, group in child_groups.items():
        for row in rows:
            child_row = [row[i] if i < len(row) else '' for i in group['child_col_indices']]
            group['child_rows'].append(child_row)

    # Extract parent rows
    parent_rows = []
    for row in rows:
        parent_row = [row[i] if i < len(row) else '' for i in parent_col_indices]
        parent_rows.append(parent_row)

    return {
        'parent_headers': parent_headers,
        'parent_rows': parent_rows,
        'child_groups': child_groups,
    }


def _parse_csv(file_obj):
    """Parse CSV — auto-detect delimiter."""
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode('utf-8-sig')

    sniffer = csv.Sniffer()
    dialect = sniffer.sniff(raw[:1024])
    reader = csv.reader(io.StringIO(raw), dialect)

    rows = list(reader)
    if not rows:
        return {'headers': [], 'rows': [], 'sheets': []}

    headers = rows[0]
    data_rows = rows[1:]

    return {
        'headers': headers,
        'rows': data_rows,
        'sheets': [],
    }


def _parse_xlsx(file_obj, sheet_name=None):
    """Parse Excel file — return sheet names + data from selected sheet."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    if sheet_name and sheet_name in sheet_names:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'headers': [], 'rows': [], 'sheets': sheet_names}

    headers = [str(h) if h is not None else '' for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append([str(v) if v is not None else '' for v in row])

    return {
        'headers': headers,
        'rows': data_rows,
        'sheets': sheet_names,
    }
